import io
import os
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import app as farm_app
import database


class DummyResponse:
    def __init__(self, payload=None, status_code=200, json_error=None):
        self.payload = payload
        self.status_code = status_code
        self._json_error = json_error

    def raise_for_status(self):
        if self.status_code >= 400:
            raise farm_app.requests.HTTPError(f'HTTP {self.status_code}')

    def json(self):
        if self._json_error is not None:
            raise self._json_error
        return self.payload


@pytest.fixture
def client(tmp_path):
    original_db_path = database.DB_PATH
    original_data_dir = database._DATA_DIR
    database._DATA_DIR = str(tmp_path / 'data')
    database.DB_PATH = str(tmp_path / 'runtime.db')
    database.init()
    farm_app._mids_cache = {}
    farm_app._mids_cache_ts = 0
    farm_app._pacifica_prices = {}
    farm_app._pacifica_prices_ts = 0
    farm_app._ethereal_products = {}
    farm_app._ethereal_products_ts = 0
    yield farm_app.app.test_client()
    database.DB_PATH = original_db_path
    database._DATA_DIR = original_data_dir


def _make_balance_workbook(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['#', 'Address', 'Volume', 'Burn', 'Points', 'PPrice', 'Balance'])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _seed_protocol_wallet(protocol='Proto', address='0x1111111111111111111111111111111111111111'):
    database.add_protocol(protocol)
    database.bulk_add_wallets([address], protocols=[protocol])
    return database.get_wallets()[0]


def test_get_json_rejects_invalid_json(monkeypatch):
    def fake_get(*args, **kwargs):
        return DummyResponse(json_error=ValueError('bad json'))

    monkeypatch.setattr(farm_app.requests, 'get', fake_get)

    with pytest.raises(RuntimeError, match='Invalid JSON'):
        farm_app._get_json('https://example.com/data')


def test_post_json_rejects_invalid_json(monkeypatch):
    def fake_post(*args, **kwargs):
        return DummyResponse(json_error=ValueError('bad json'))

    monkeypatch.setattr(farm_app.requests, 'post', fake_post)

    with pytest.raises(RuntimeError, match='Invalid JSON'):
        farm_app._post_json('https://example.com/data')


def test_profile_and_index_pages_render(client):
    assert client.get('/profiles').status_code == 200
    assert client.get('/').status_code == 200


def test_hl_mids_returns_502_on_upstream_http_error(monkeypatch, client):
    def fake_post(*args, **kwargs):
        return DummyResponse(status_code=503)

    monkeypatch.setattr(farm_app.requests, 'post', fake_post)
    response = client.get('/api/hl/mids')

    assert response.status_code == 502
    assert 'HTTP 503' in response.get_json()['error']


def test_select_profile_rejects_invalid_name(client):
    response = client.post('/api/profiles/select', json={'name': '../bad'})

    assert response.status_code == 400
    assert response.get_json()['error'] == 'Недопустимое имя'


def test_rename_active_profile_updates_current_profile(client):
    assert client.post('/api/profiles/select', json={'name': 'alpha'}).status_code == 200

    response = client.post('/api/profiles/rename', json={'old_name': 'alpha', 'new_name': 'beta'})

    assert response.status_code == 200
    assert client.get('/api/profiles/current').get_json()['name'] == 'beta'
    assert client.get('/api/profiles').get_json() == ['beta']


def test_delete_active_profile_is_rejected(client):
    client.post('/api/profiles/select', json={'name': 'active'})

    response = client.post('/api/profiles/delete', json={'name': 'active'})

    assert response.status_code == 400
    assert response.get_json()['error'] == 'Cannot delete the active profile'


def test_rename_profile_rejects_invalid_and_taken_names(client):
    client.post('/api/profiles/select', json={'name': 'alpha'})
    client.post('/api/profiles/select', json={'name': 'beta'})

    bad = client.post('/api/profiles/rename', json={'old_name': 'alpha', 'new_name': '../bad'})
    taken = client.post('/api/profiles/rename', json={'old_name': 'alpha', 'new_name': 'beta'})

    assert bad.status_code == 400
    assert taken.status_code == 400
    assert taken.get_json()['error'] == 'Такое имя уже занято'


def test_rename_profile_requires_both_names_and_existing_old_profile(client):
    missing = client.post('/api/profiles/rename', json={'old_name': '', 'new_name': ''})
    not_found = client.post('/api/profiles/rename', json={'old_name': 'ghost', 'new_name': 'new'})

    assert missing.status_code == 400
    assert missing.get_json()['error'] == 'Укажите оба имени'
    assert not_found.status_code == 404


def test_delete_profile_rejects_missing_and_invalid_names(client):
    missing = client.post('/api/profiles/delete', json={'name': 'ghost'})
    invalid = client.post('/api/profiles/delete', json={'name': '../ghost'})

    assert missing.status_code == 404
    assert invalid.status_code == 400


def test_add_protocol_requires_name(client):
    response = client.post('/api/protocols', json={'name': '   '})

    assert response.status_code == 400
    assert response.get_json()['error'] == 'Название обязательно'


def test_stats_route_returns_expected_totals(client):
    database.add_protocol('Proto', spent=10, earned=25, status='фармлю')

    response = client.get('/api/stats')

    assert response.status_code == 200
    payload = response.get_json()
    assert payload['total'] == 1
    assert payload['active'] == 1
    assert payload['pnl'] == 15


def test_protocols_endpoint_uses_manual_values_without_wallet_data(client):
    database.add_protocol('Proto', spent=12, earned=30)

    response = client.get('/api/protocols')

    assert response.status_code == 200
    payload = response.get_json()[0]
    assert payload['total_spent'] == 12
    assert payload['pnl'] == 18


def test_add_protocol_rejects_duplicate_name(client):
    assert client.post('/api/protocols', json={'name': 'Proto'}).status_code == 200

    response = client.post('/api/protocols', json={'name': 'Proto'})

    assert response.status_code == 400


def test_protocols_endpoint_uses_wallet_aggregates_over_manual_values(client):
    database.add_protocol('Proto', spent=999, earned=10)
    database.bulk_add_wallets(['0x1111111111111111111111111111111111111111'], protocols=['Proto'])
    wallet = database.get_wallets()[0]
    database.update_wallet_protocol(wallet['id'], 'Proto', deposit=100, wallet_balance=65, wp_earned=20)

    response = client.get('/api/protocols')

    assert response.status_code == 200
    payload = response.get_json()
    assert payload[0]['total_spent'] == 15
    assert payload[0]['pnl'] == -15


def test_patch_protocol_returns_404_for_missing_id(client):
    response = client.patch('/api/protocols/999', json={'spent': 10})

    assert response.status_code == 404
    assert response.get_json()['error'] == 'Not found'


def test_patch_protocol_preserves_unspecified_fields(client):
    database.add_protocol('Proto', spent=10, earned=20, status='фармлю', note='old', color='#111111', points=5)
    pid = database.get_protocols()[0]['id']

    response = client.patch('/api/protocols/{0}'.format(pid), json={'status': 'done'})

    assert response.status_code == 200
    updated = database.get_protocols()[0]
    assert updated['status'] == 'done'
    assert updated['note'] == 'old'
    assert updated['color'] == '#111111'
    assert updated['points'] == 5


def test_put_protocol_updates_all_fields(client):
    database.add_protocol('Proto', spent=1, earned=2, status='old', note='old', color='#111111', points=1)
    pid = database.get_protocols()[0]['id']

    response = client.put(f'/api/protocols/{pid}', json={
        'spent': 25,
        'earned': 40,
        'status': 'done',
        'note': 'new',
        'color': '#222222',
        'points': 9,
        'point_price': 0.75,
    })

    assert response.status_code == 200
    updated = database.get_protocols()[0]
    assert updated['spent'] == 25
    assert updated['earned'] == 40
    assert updated['status'] == 'done'
    assert updated['note'] == 'new'
    assert updated['color'] == '#222222'
    assert updated['points'] == 9
    assert updated['point_price'] == 0.75


def test_delete_protocol_via_api_removes_wallet_links(client):
    wallet = _seed_protocol_wallet()
    pid = database.get_protocols()[0]['id']

    response = client.delete(f'/api/protocols/{pid}')

    assert response.status_code == 200
    assert database.get_protocols() == []
    assert database.get_wallets()[0]['protocols'] == []


def test_protocol_wallets_and_import_routes_work(client):
    _seed_protocol_wallet()

    imported = client.post('/api/protocols/Proto/import', json={
        'rows': [{'address': '0x1111111111111111111111111111111111111111', 'wallet_balance': 77, 'wp_points': 5}],
        'add_points': False,
        'add_deposit': False,
    })
    details = client.get('/api/protocols/Proto/wallets')

    assert imported.status_code == 200
    assert imported.get_json()['matched'] == 1
    assert details.status_code == 200
    assert details.get_json()[0]['wallet_balance'] == 77


def test_week_routes_cover_create_patch_get_wallets_and_delete(client):
    wallet = _seed_protocol_wallet()
    database.import_protocol_wallet_data('Proto', [{
        'address': wallet['address'],
        'deposit': 100,
        'wallet_balance': 80,
        'wp_points': 10,
        'wp_earned': 5,
    }], week='2026-05-01')

    created = client.post('/api/protocols/Proto/weeks', json={'week_start': '2026-05-01', 'week_end': '2026-05-07'})
    wid = created.get_json()['id']
    listing = client.get('/api/protocols/Proto/weeks')
    wallets = client.get('/api/protocols/Proto/weeks/2026-05-01/wallets')
    patched = client.patch(f'/api/protocols/Proto/weeks/{wid}', json={'week_start': '2026-05-02', 'week_end': '2026-05-08'})
    deleted = client.delete(f'/api/protocols/Proto/weeks/{wid}')

    assert created.status_code == 200
    assert listing.status_code == 200
    assert listing.get_json()[0]['week_start'] == '2026-05-01'
    assert wallets.status_code == 200
    assert wallets.get_json()[0]['deposit_this_week'] == 100
    assert patched.status_code == 200
    assert deleted.status_code == 200
    assert client.get('/api/protocols/Proto/weeks').get_json() == []


def test_patch_week_requires_dates(client):
    response = client.patch('/api/protocols/Proto/weeks/1', json={'week_start': '', 'week_end': ''})

    assert response.status_code == 400
    assert response.get_json()['error'] == 'Укажите даты'


def test_clear_protocol_data_route_zeroes_financials(client):
    wallet = _seed_protocol_wallet()
    database.update_wallet_protocol(wallet['id'], 'Proto', deposit=100, wallet_balance=50, wp_points=9, wp_earned=7)

    response = client.post('/api/protocols/Proto/clear-data')

    assert response.status_code == 200
    row = database.get_protocol_wallets_detail('Proto')[0]
    assert row['deposit'] == 0
    assert row['wallet_balance'] == 0
    assert row['wp_points'] == 0
    assert row['wp_earned'] == 0


def test_wallet_protocol_patch_updates_selected_fields(client):
    wallet = _seed_protocol_wallet()

    response = client.patch('/api/wallet-protocols', json={
        'wallet_id': wallet['id'],
        'protocol': 'Proto',
        'deposit': 10,
        'wallet_balance': 8,
        'wp_points': 3,
        'wp_earned': 1,
    })

    assert response.status_code == 200
    row = database.get_protocol_wallets_detail('Proto')[0]
    assert row['deposit'] == 10
    assert row['wallet_balance'] == 8
    assert row['wp_points'] == 3
    assert row['wp_earned'] == 1


def test_wallet_filters_route_supports_protocol_unassigned_and_chain(client):
    database.add_protocol('Proto')
    database.bulk_add_wallets([
        '0x1111111111111111111111111111111111111111',
        'So11111111111111111111111111111111111111112',
    ], protocols=['Proto'])
    database.bulk_add_wallets(['0x2222222222222222222222222222222222222222'])

    by_protocol = client.get('/api/wallets?protocol=Proto')
    unassigned = client.get('/api/wallets?unassigned=1')
    by_chain = client.get('/api/wallets?chain=sol')

    assert len(by_protocol.get_json()) == 2
    assert len(unassigned.get_json()) == 1
    assert by_chain.get_json()[0]['chain'] == 'sol'


def test_add_wallet_requires_protocol_and_address(client):
    response = client.post('/api/wallets', json={'protocol': '', 'address': ''})

    assert response.status_code == 400
    assert response.get_json()['error'] == 'Протокол и адрес обязательны'


def test_add_wallet_delete_wallet_and_bulk_delete_routes(client):
    database.add_protocol('Proto')
    added = client.post('/api/wallets', json={'protocol': 'Proto', 'address': '0x1111111111111111111111111111111111111111', 'label': 'x'})
    wid = database.get_wallets()[0]['id']
    database.bulk_add_wallets(['0x2222222222222222222222222222222222222222'])
    wid2 = [w for w in database.get_wallets() if w['id'] != wid][0]['id']
    bulk_deleted = client.post('/api/wallets/bulk-delete', json={'ids': [wid2]})
    deleted = client.delete(f'/api/wallets/{wid}')

    assert added.status_code == 200
    assert bulk_deleted.get_json()['count'] == 1
    assert deleted.status_code == 200
    assert database.get_wallets() == []


def test_bulk_add_wallets_requires_non_empty_addresses(client):
    response = client.post('/api/wallets/bulk', json={'addresses': ' \n  '})

    assert response.status_code == 400
    assert response.get_json()['error'] == 'Нет адресов'


def test_bulk_add_wallets_accepts_newline_string_and_assigns_protocols(client):
    database.add_protocol('P1')
    response = client.post('/api/wallets/bulk', json={
        'addresses': '0x1111111111111111111111111111111111111111\n0x2222222222222222222222222222222222222222',
        'protocols': ['P1'],
        'label': 'seed',
    })

    assert response.status_code == 200
    assert response.get_json()['count'] == 2
    wallets = database.get_wallets(protocol='P1')
    assert len(wallets) == 2
    assert all(w['label'] == 'seed' for w in wallets)


def test_assign_wallets_requires_ids(client):
    response = client.put('/api/wallets/assign', json={'ids': [], 'protocol': 'P1'})

    assert response.status_code == 400
    assert response.get_json()['error'] == 'Нет кошельков'


def test_assign_wallets_bulk_adds_protocol_without_replacing_existing(client):
    database.add_protocol('P1')
    database.add_protocol('P2')
    database.bulk_add_wallets(['0x1111111111111111111111111111111111111111'], protocols=['P1'])
    wid = database.get_wallets()[0]['id']

    response = client.put('/api/wallets/assign', json={'ids': [wid], 'protocol': 'P2'})

    assert response.status_code == 200
    assert database.get_wallets()[0]['protocols'] == ['P1', 'P2']


def test_assign_wallets_with_protocols_replaces_existing_links(client):
    database.add_protocol('P1')
    database.add_protocol('P2')
    database.bulk_add_wallets(['0x1111111111111111111111111111111111111111'], protocols=['P1'])
    wid = database.get_wallets()[0]['id']

    response = client.put('/api/wallets/assign', json={'ids': [wid], 'protocols': ['P2']})

    assert response.status_code == 200
    wallet = database.get_wallets()[0]
    assert wallet['protocols'] == ['P2']


def test_update_wallet_route_updates_chain_and_proxy(client):
    database.bulk_add_wallets(['0x1111111111111111111111111111111111111111'])
    wid = database.get_wallets()[0]['id']

    response = client.patch(f'/api/wallets/{wid}', json={'chain': 'sol', 'proxy': 'http://proxy:1'})

    assert response.status_code == 200
    wallet = database.get_wallets()[0]
    assert wallet['chain'] == 'sol'
    assert wallet['proxy'] == 'http://proxy:1'


def test_import_balances_requires_file(client):
    response = client.post('/api/wallets/import-balances', data={}, content_type='multipart/form-data')

    assert response.status_code == 400
    assert response.get_json()['error'] == 'Нет файла'


def test_import_balances_reads_xlsx_and_reports_counts(client):
    address = '0x1111111111111111111111111111111111111111'
    database.bulk_add_wallets([address])
    workbook = _make_balance_workbook([
        [1, '0x1111..1111', 5, 2, 100, 0.5, 42],
        [2, 'not-short', 0, 0, 0, 0, 0],
    ])

    response = client.post(
        '/api/wallets/import-balances',
        data={'file': (workbook, 'balances.xlsx')},
        content_type='multipart/form-data',
    )

    assert response.status_code == 200
    assert response.get_json()['matched'] == 1
    assert response.get_json()['total'] == 2
    wallet = database.get_wallets()[0]
    assert wallet['balance'] == 42
    assert wallet['points'] == 100


def test_import_balances_reads_multiple_sheets_and_skips_blank_rows(client):
    database.bulk_add_wallets([
        '0x1111111111111111111111111111111111111111',
        '0x2222222222222222222222222222222222222222',
    ])
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.append(['#', 'Address', 'Volume', 'Burn', 'Points', 'PPrice', 'Balance'])
    ws1.append([1, '0x1111..1111', 1, 1, 1, 1, 11])
    ws1.append([2, None, 0, 0, 0, 0, 0])
    ws2 = wb.create_sheet('Sheet2')
    ws2.append(['#', 'Address', 'Volume', 'Burn', 'Points', 'PPrice', 'Balance'])
    ws2.append([1, '0x2222..2222', 2, 2, 2, 2, 22])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = client.post('/api/wallets/import-balances', data={'file': (buf, 'multi.xlsx')}, content_type='multipart/form-data')

    assert response.status_code == 200
    assert response.get_json()['matched'] == 2
    assert response.get_json()['total'] == 2


def test_import_balances_partial_match_multiple_wallets(client):
    database.bulk_add_wallets([
        '0x111111111111111111111111111111111111aaaa',
        '0x111111111111111111111111111111111111bbbb',
    ])
    workbook = _make_balance_workbook([
        [1, '0x1111..aaaa', 1, 1, 1, 1, 11],
        [2, '0x1111..bbbb', 2, 2, 2, 2, 22],
    ])

    response = client.post('/api/wallets/import-balances', data={'file': (workbook, 'balances.xlsx')}, content_type='multipart/form-data')

    assert response.status_code == 200
    assert response.get_json()['matched'] == 2


def test_add_reminder_requires_protocol_and_date(client):
    response = client.post('/api/reminders', json={'protocol': '', 'remind_at': ''})

    assert response.status_code == 400
    assert response.get_json()['error'] == 'Протокол и дата обязательны'


def test_reminder_routes_cover_list_done_delete_and_due(client):
    created = client.post('/api/reminders', json={'protocol': 'Proto', 'remind_at': '2000-01-01T00:00'})
    rid = database.get_reminders(include_done=True)[0]['id']
    due = client.get('/api/due-reminders')
    done = client.post(f'/api/reminders/{rid}/done')
    listed = client.get('/api/reminders?include_done=1')
    deleted = client.delete(f'/api/reminders/{rid}')

    assert created.status_code == 200
    assert len(due.get_json()) == 1
    assert done.status_code == 200
    assert listed.get_json()[0]['done'] == 1
    assert deleted.status_code == 200
    assert client.get('/api/reminders?include_done=1').get_json() == []


def test_extended_all_positions_keeps_partial_failures(client, monkeypatch):
    database.add_extended_account('ok', 'key-ok')
    database.add_extended_account('bad', 'key-bad')

    def fake_fetch(api_key):
        if api_key == 'key-bad':
            raise RuntimeError('boom')
        return {'positions': [{'market': 'BTC'}], 'equity': 10, 'available': 5, 'unrealised_pnl': 1}

    monkeypatch.setattr(farm_app, '_fetch_extended_account', fake_fetch)

    response = client.get('/api/extended/all-positions')

    assert response.status_code == 200
    accounts = response.get_json()['accounts']
    assert len(accounts) == 2
    assert any(acc['label'] == 'ok' and acc['equity'] == 10 for acc in accounts)
    assert any(acc['label'] == 'bad' and acc['error'] == 'boom' for acc in accounts)


def test_fetch_extended_account_parses_payload(monkeypatch):
    responses = iter([
        {'data': {'equity': '123.45', 'availableForTrade': '100', 'unrealisedPnl': '5.5'}},
        {'data': [{
            'size': '2',
            'openPrice': '10',
            'markPrice': '12',
            'liquidationPrice': '8',
            'unrealisedPnl': '4',
            'margin': '20',
            'leverage': {'value': '3'},
            'side': 'LONG',
            'market': 'BTC',
        }]}
    ])
    monkeypatch.setattr(farm_app, '_get_json', lambda *args, **kwargs: next(responses))

    result = farm_app._fetch_extended_account('key')

    assert result['equity'] == 123.45
    assert result['available'] == 100.0
    assert result['margin_used'] == 20.0
    assert result['margin_usage'] == 16.2
    assert result['account_health'] == 81.0
    assert result['positions'][0]['market'] == 'BTC'
    assert result['positions'][0]['dist_liq'] == 33.33


def test_extended_account_routes_cover_crud_and_group_patch(client):
    bad = client.post('/api/extended/accounts', json={'label': '', 'api_key': ''})
    created = client.post('/api/extended/accounts', json={'label': 'acc', 'api_key': 'key1'})
    aid = database.get_extended_accounts()[0]['id']
    patched = client.patch(f'/api/extended/accounts/{aid}', json={'group_name': 'grp'})
    listed = client.get('/api/extended/accounts')
    deleted = client.delete(f'/api/extended/accounts/{aid}')

    assert bad.status_code == 400
    assert created.status_code == 200
    assert patched.status_code == 200
    assert listed.get_json()[0]['group_name'] == 'grp'
    assert deleted.status_code == 200


def test_extended_all_positions_empty_accounts(client):
    response = client.get('/api/extended/all-positions')

    assert response.status_code == 200
    assert response.get_json() == {'accounts': []}


def test_ethereal_products_parse_public_payload(monkeypatch):
    farm_app._ethereal_products = {}
    farm_app._ethereal_products_ts = 0
    monkeypatch.setattr(farm_app, '_get_json', lambda *args, **kwargs: {
        'data': [{'id': 'product-1', 'ticker': 'ETHUSD', 'displayTicker': 'ETH-USD'}]
    })

    assert farm_app._ethereal_get_products() == {'product-1': 'ETH-USD'}


def test_fetch_ethereal_account_parses_public_payload(monkeypatch):
    monkeypatch.setattr(farm_app, '_ethereal_get_products', lambda: {'product-1': 'BTC-USD'})

    def fake_get_json(url, **kwargs):
        if url.endswith('/v1/subaccount'):
            return {'data': [{'id': 'sub-1'}], 'hasNext': False}
        if url.endswith('/v1/position'):
            return {'data': [{
                'productId': 'product-1',
                'side': 0,
                'size': '0.002',
                'cost': '150',
                'unrealizedPnl': '10.5',
            }], 'hasNext': False}
        if url.endswith('/v1/subaccount/balance'):
            return {'data': [{
                'tokenName': 'USD',
                'amount': '42',
                'available': '30',
                'totalUsed': '10',
            }], 'hasNext': False}
        if url.endswith('/v1/product/market-price'):
            assert kwargs['params'] == [('productIds[]', 'product-1')]
            return {'data': [{'productId': 'product-1', 'oraclePrice': '80000'}]}
        raise AssertionError(url)

    monkeypatch.setattr(farm_app, '_get_json', fake_get_json)

    result = farm_app._fetch_ethereal_account('0xabc')
    position = result['positions'][0]

    assert result['equity'] == 42.0
    assert result['available'] == 30.0
    assert result['margin_used'] == 10.0
    assert result['unrealised_pnl'] == 10.5
    assert position['symbol'] == 'BTC-USD'
    assert position['direction'] == 'long'
    assert position['entry_price'] == 75000.0
    assert position['current_price'] == 80000.0
    assert position['unrealized_pnl'] == 10.5
    assert position['margin'] == 10.0
    assert position['leverage'] == 15.0
    assert result['margin_usage'] == 23.81
    assert result['account_health'] == 71.43


def test_pacifica_account_routes_and_partial_failures(client, monkeypatch):
    bad = client.post('/api/pacifica/accounts', json={'label': '', 'address': ''})
    client.post('/api/pacifica/accounts', json={'label': 'ok', 'address': '0x1'})
    client.post('/api/pacifica/accounts', json={'label': 'bad', 'address': '0x2'})
    accounts = database.get_pacifica_accounts()
    ok_id = next(a['id'] for a in accounts if a['label'] == 'ok')

    patched = client.patch(f'/api/pacifica/accounts/{ok_id}', json={'group_name': 'grp'})

    def fake_fetch(address):
        if address == '0x2':
            raise RuntimeError('pacifica down')
        return {'positions': [], 'equity': 5, 'available': 2, 'margin_used': 1, 'maint_margin': 0, 'unrealised_pnl': 0, 'account_leverage': None, 'margin_ratio': None}

    monkeypatch.setattr(farm_app, '_fetch_pacifica_account', fake_fetch)
    response = client.get('/api/pacifica/all-positions')

    assert bad.status_code == 400
    assert patched.status_code == 200
    assert len(response.get_json()['accounts']) == 2
    assert any(acc['label'] == 'ok' and acc['group_name'] == 'grp' for acc in response.get_json()['accounts'])
    assert any(acc['label'] == 'bad' and acc['error'] == 'pacifica down' for acc in response.get_json()['accounts'])


def test_fetch_pacifica_account_parses_payload(monkeypatch):
    monkeypatch.setattr(farm_app, '_pacifica_get_prices', lambda: {'BTC': 12})
    responses = iter([
        {'data': [{'symbol': 'BTC', 'side': 'bid', 'amount': '2', 'entry_price': '10', 'funding': '1', 'isolated': False, 'liq_price': '7'}]},
        {'data': {'total_margin_used': '40', 'account_equity': '100', 'cross_mmr': '10', 'balance': '90', 'available_to_spend': '50'}},
    ])
    monkeypatch.setattr(farm_app, '_get_json', lambda *args, **kwargs: next(responses))

    result = farm_app._fetch_pacifica_account('0x1')

    assert result['equity'] == 100.0
    assert result['positions'][0]['unrealised_pnl'] == 5.0
    assert result['positions'][0]['leverage'] == 0.6
    assert result['margin_ratio'] == 10.0
    assert result['margin_usage'] == 40.0
    assert result['account_health'] == 50.0


def test_nado_account_routes_and_partial_failures(client, monkeypatch):
    bad = client.post('/api/nado/accounts', json={'label': '', 'address': ''})
    client.post('/api/nado/accounts', json={'label': 'ok', 'address': '0x1'})
    client.post('/api/nado/accounts', json={'label': 'bad', 'address': '0x2'})
    accounts = database.get_nado_accounts()
    ok_id = next(a['id'] for a in accounts if a['label'] == 'ok')
    client.patch(f'/api/nado/accounts/{ok_id}', json={'group_name': 'grp'})

    def fake_fetch(address):
        if address == '0x2':
            raise RuntimeError('nado down')
        return {'positions': [], 'subaccount': 'x', 'account_value': 1, 'available_margin': 2}

    monkeypatch.setattr(farm_app, '_fetch_nado_account', fake_fetch)
    response = client.get('/api/nado/all-positions')

    assert bad.status_code == 400
    assert len(response.get_json()['accounts']) == 2
    assert any(acc['label'] == 'bad' and acc['error'] == 'nado down' for acc in response.get_json()['accounts'])


def test_fetch_nado_account_parses_payload_and_tolerates_missing_info(monkeypatch):
    monkeypatch.setattr(farm_app, '_nado_get_symbols', lambda: {'1': 'BTC'})
    responses = iter([
        {'status': 'success', 'data': {'isolated_positions': [{
            'base_balance': {'balance': {'amount': str(2 * 10**18), 'v_quote_balance': str(20 * 10**18)}, 'product_id': '1'},
            'quote_balance': {'balance': {'amount': str(5 * 10**18)}},
            'base_product': {'oracle_price_x18': str(12 * 10**18), 'risk': {'long_weight_maintenance_x18': str(1 * 10**18)}},
            'healths': [{}, {'assets': str(100 * 10**18), 'health': str(20 * 10**18)}],
        }]}},
        {'data': {'healths': [{'assets': str(30 * 10**18), 'health': str(10 * 10**18)}]}},
    ])
    monkeypatch.setattr(farm_app, '_get_json', lambda *args, **kwargs: next(responses))

    result = farm_app._fetch_nado_account('0x1')

    assert result['available_margin'] == 10.0
    assert result['account_value'] == 19.0
    assert result['margin_used'] == 5.0
    assert result['margin_usage'] == 26.32
    assert result['account_health'] == 52.63
    assert result['account_leverage'] == 1.26
    assert result['positions'][0]['symbol'] == 'BTC'
    assert result['positions'][0]['unrealized_pnl'] == 4.0


def test_nado_get_symbols_falls_back_to_cache_on_error(monkeypatch):
    farm_app._nado_symbols = {'1': 'BTC'}
    farm_app._nado_symbols_ts = 0
    monkeypatch.setattr(farm_app, '_get_json', lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError('boom')))

    assert farm_app._nado_get_symbols() == {'1': 'BTC'}


def test_hl_account_routes_and_position_endpoints(client, monkeypatch):
    bad = client.post('/api/hl/accounts', json={'label': '', 'address': ''})
    client.post('/api/hl/accounts', json={'label': 'ok', 'address': '0x1'})
    client.post('/api/hl/accounts', json={'label': 'bad', 'address': '0x2'})
    accounts = database.get_hl_accounts()
    ok_id = next(a['id'] for a in accounts if a['label'] == 'ok')
    client.patch(f'/api/hl/accounts/{ok_id}', json={'group_name': 'grp'})

    def fake_fetch(address, mids):
        if address == '0x2':
            raise RuntimeError('hl down')
        return {'positions': [{'coin': 'BTC'}], 'summary': {'account_value': 1}}

    monkeypatch.setattr(farm_app, '_fetch_hl_account', fake_fetch)
    monkeypatch.setattr(farm_app, '_post_json', lambda *args, **kwargs: {'BTC': '100'})
    response = client.get('/api/hl/all-positions')
    mids = client.get('/api/hl/mids')

    assert bad.status_code == 400
    assert response.status_code == 200
    assert any(acc['label'] == 'bad' and acc['error'] == 'hl down' for acc in response.get_json()['accounts'])
    assert mids.get_json() == {'BTC': '100'}


def test_fetch_hl_account_parses_payload_with_mock(monkeypatch):
    state = {
        'assetPositions': [{
            'position': {
                'szi': '2',
                'coin': 'BTC',
                'liquidationPx': '8',
                'entryPx': '10',
                'marginUsed': '5',
                'positionValue': '24',
                'unrealizedPnl': '4',
                'leverage': {'value': '3', 'type': 'cross'},
                'returnOnEquity': '0.2',
            }
        }],
        'marginSummary': {'accountValue': '100', 'totalMarginUsed': '5', 'totalUnrealizedPnl': '4'},
        'withdrawable': '80',
    }
    monkeypatch.setattr(farm_app, '_post_json', lambda *args, **kwargs: state)

    result = farm_app._fetch_hl_account('0x1', {'BTC': '12'})

    assert result['positions'][0]['coin'] == 'BTC'
    assert result['positions'][0]['dist_liq'] == 33.33
    assert result['summary']['account_value'] == 100.0
    assert result['summary']['margin_usage'] == 5.0
    assert result['summary']['account_health'] == 80.0
    assert result['summary']['account_leverage'] == 0.24


def test_hl_all_positions_returns_502_when_mids_fetch_fails(client, monkeypatch):
    database.add_hl_account('acc', '0x1')
    monkeypatch.setattr(farm_app, '_post_json', lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError('mid fail')))
    farm_app._mids_cache = {}
    farm_app._mids_cache_ts = 0

    response = client.get('/api/hl/all-positions')

    assert response.status_code == 502
    assert 'mid fail' in response.get_json()['error']


def test_hl_search_uses_cached_mids_and_filters(client):
    farm_app._mids_cache = {'BTC': '100000', 'ETH': '2000', 'DOGE': '0.1'}
    farm_app._mids_cache_ts = farm_app.time.time()

    response = client.get('/api/hl/search?q=et')

    assert response.status_code == 200
    assert response.get_json() == [{'symbol': 'ETH', 'price': 2000.0}]


def test_positions_and_groups_routes_cover_crud_and_validation(client):
    bad_group = client.post('/api/position-groups', json={'name': '  '})
    good_group = client.post('/api/position-groups', json={'name': 'Main'})
    gid = good_group.get_json()['id']
    bad_position = client.post('/api/positions', json={'symbol': 'BTC'})
    good_position = client.post('/api/positions', json={
        'symbol': 'btc',
        'direction': 'long',
        'collateral': 100,
        'leverage': 5,
        'entry_price': 25000,
        'note': 'n1',
        'group_id': gid,
    })
    pid = database.get_positions()[0]['id']
    listed = client.get('/api/positions')
    updated = client.patch(f'/api/positions/{pid}', json={
        'symbol': 'eth',
        'direction': 'short',
        'collateral': 50,
        'leverage': 2,
        'entry_price': 3000,
        'note': 'n2',
        'group_id': None,
    })
    renamed = client.patch(f'/api/position-groups/{gid}', json={'name': 'Renamed'})
    deleted_group = client.delete(f'/api/position-groups/{gid}')
    deleted_position = client.delete(f'/api/positions/{pid}')

    assert bad_group.status_code == 400
    assert bad_position.status_code == 400
    assert good_position.status_code == 200
    assert listed.get_json()[0]['symbol'] == 'BTC'
    assert updated.status_code == 200
    assert renamed.status_code == 200
    assert deleted_group.status_code == 200
    assert deleted_position.status_code == 200


def test_rename_position_group_requires_name(client):
    gid = client.post('/api/position-groups', json={'name': 'Main'}).get_json()['id']

    response = client.patch(f'/api/position-groups/{gid}', json={'name': '   '})

    assert response.status_code == 400


def test_proxies_routes_cover_bulk_list_assign_unassign_delete(client):
    database.bulk_add_wallets(['0x1111111111111111111111111111111111111111'])
    wid = database.get_wallets()[0]['id']

    added = client.post('/api/proxies/bulk', json={'proxies': ['http://proxy:1', 'http://proxy:2']})
    proxies = client.get('/api/proxies').get_json()
    pid = proxies[0]['id']
    patched_assign = client.patch(f'/api/proxies/{pid}', json={'wallet_id': wid, 'label': 'A'})
    listed = client.get('/api/proxies')
    patched_unassign = client.patch(f'/api/proxies/{pid}', json={'wallet_id': None})
    deleted = client.delete(f'/api/proxies/{pid}')

    assert added.status_code == 200
    assert added.get_json()['added'] == 2
    assert patched_assign.status_code == 200
    assert any(p['label'] == 'A' for p in listed.get_json())
    assert patched_unassign.status_code == 200
    assert deleted.status_code == 200


def test_open_folder_route_calls_explorer_only_for_existing_path(client, monkeypatch, tmp_path):
    calls = []
    monkeypatch.setattr('subprocess.Popen', lambda args: calls.append(args))
    existing = tmp_path / 'file.txt'
    existing.write_text('x', encoding='utf-8')

    ok = client.post('/api/open-folder', json={'path': str(existing)})
    missing = client.post('/api/open-folder', json={'path': str(tmp_path / 'missing.txt')})

    assert ok.status_code == 200
    assert missing.status_code == 200
    assert len(calls) == 1
    assert calls[0][0] == 'explorer'


def test_export_route_writes_workbook_with_expected_sheets(client, monkeypatch, tmp_path):
    wallet = _seed_protocol_wallet()
    database.update_wallet_protocol(wallet['id'], 'Proto', deposit=100, wallet_balance=60, wp_earned=20)
    downloads = tmp_path / 'Downloads'
    monkeypatch.setattr(farm_app.os.path, 'expanduser', lambda _: str(tmp_path))

    response = client.get('/api/export')

    assert response.status_code == 200
    payload = response.get_json()
    exported = Path(payload['path'])
    assert exported.exists()
    wb = openpyxl.load_workbook(exported)
    assert wb.sheetnames == ['Протоколы', 'Кошельки', 'Сводка']
    assert wb['Протоколы']['A2'].value == 'Proto'
    assert wb['Кошельки']['A2'].value == wallet['address']
    assert wb['Сводка']['A2'].value == 'Всего протоколов'


def test_export_route_marks_positive_and_negative_pnl_colors(client, monkeypatch, tmp_path):
    database.add_protocol('Win', spent=10, earned=20)
    database.add_protocol('Loss', spent=20, earned=10)
    monkeypatch.setattr(farm_app.os.path, 'expanduser', lambda _: str(tmp_path))

    payload = client.get('/api/export').get_json()
    wb = openpyxl.load_workbook(payload['path'])

    assert wb['Протоколы']['D2'].font.color.rgb.endswith('1D9E75')
    assert wb['Протоколы']['D3'].font.color.rgb.endswith('E24B4A')
