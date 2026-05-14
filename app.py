import os
import io
import time
import requests
from datetime import datetime
from flask import Flask, jsonify, request, render_template, send_file
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import database as db
import telegram_bot as tg

_price_cache: dict = {}
_price_cache_ts: float = 0
_PRICE_TTL = 30  # seconds

_ethereal_products: dict = {}
_ethereal_products_ts: float = 0

app = Flask(__name__)

with app.app_context():
    db.init()

_tg_monitor: tg.TelegramMonitor | None = None


def init_telegram_monitor() -> None:
    global _tg_monitor
    cfg = db.get_telegram_config()
    _tg_monitor = tg.TelegramMonitor()
    _tg_monitor.update_config(cfg)
    if cfg.get('enabled') and cfg.get('bot_token') and cfg.get('chat_id'):
        _tg_monitor.start()


def _get_json(url: str, **kwargs):
    resp = requests.get(url, **kwargs)
    resp.raise_for_status()
    try:
        return resp.json()
    except ValueError as exc:
        raise RuntimeError(f'Invalid JSON from {url}') from exc


def _post_json(url: str, **kwargs):
    resp = requests.post(url, **kwargs)
    resp.raise_for_status()
    try:
        return resp.json()
    except ValueError as exc:
        raise RuntimeError(f'Invalid JSON from {url}') from exc


def _account_risk_metrics(equity, margin_used, available=None) -> dict:
    equity = float(equity or 0)
    margin_used = float(margin_used or 0)
    if equity <= 0:
        return {'margin_usage': None, 'account_health': None}

    margin_usage = round(margin_used / equity * 100, 2)
    if available is not None:
        health = round(float(available or 0) / equity * 100, 2)
    else:
        health = round(100 - margin_usage, 2)
    health = max(0.0, min(100.0, health))
    return {'margin_usage': margin_usage, 'account_health': health}


@app.route('/profiles')
def profiles_page():
    return render_template('profiles.html')


@app.route('/api/profiles', methods=['GET'])
def get_profiles():
    return jsonify(db.list_profiles())


@app.route('/api/profiles/select', methods=['POST'])
def select_profile():
    name = (request.json.get('name') or '').strip()
    if not name or '/' in name or '\\' in name or '..' in name:
        return jsonify({'error': 'Недопустимое имя'}), 400
    db.set_profile(name)
    return jsonify({'ok': True})


@app.route('/api/profiles/current', methods=['GET'])
def current_profile():
    import os
    name = os.path.splitext(os.path.basename(db.DB_PATH))[0]
    return jsonify({'name': name})


@app.route('/api/profiles/rename', methods=['POST'])
def rename_profile():
    import os
    old_name = (request.json.get('old_name') or '').strip()
    new_name = (request.json.get('new_name') or '').strip()
    if not old_name or not new_name:
        return jsonify({'error': 'Укажите оба имени'}), 400
    for n in (old_name, new_name):
        if '/' in n or '\\' in n or '..' in n:
            return jsonify({'error': 'Недопустимые символы'}), 400
    old_path = os.path.join(db._DATA_DIR, f'{old_name}.db')
    new_path = os.path.join(db._DATA_DIR, f'{new_name}.db')
    if not os.path.exists(old_path):
        return jsonify({'error': 'Профиль не найден'}), 404
    if os.path.exists(new_path):
        return jsonify({'error': 'Такое имя уже занято'}), 400
    os.rename(old_path, new_path)
    if db.DB_PATH == old_path:
        db.DB_PATH = new_path
    return jsonify({'ok': True})


@app.route('/api/profiles/delete', methods=['POST'])
def delete_profile():
    import os
    name = (request.json.get('name') or '').strip()
    if not name or '/' in name or '\\' in name or '..' in name:
        return jsonify({'error': 'Invalid name'}), 400
    path = os.path.join(db._DATA_DIR, f'{name}.db')
    if not os.path.exists(path):
        return jsonify({'error': 'Profile not found'}), 404
    if db.DB_PATH == path:
        return jsonify({'error': 'Cannot delete the active profile'}), 400
    os.remove(path)
    return jsonify({'ok': True})


@app.route('/')
def index():
    return render_template('index.html')


# --- Stats ---

@app.route('/api/stats')
def stats():
    s = db.get_stats()
    s['pnl'] = round(s['total_earned'] - s['total_spent'], 2)
    return jsonify(s)


# --- Protocols ---

@app.route('/api/protocols', methods=['GET'])
def get_protocols():
    protocols = db.get_protocols()
    for p in protocols:
        deposit = p.get('total_deposit') or 0
        balance = p.get('total_balance') or 0
        wp_earned = p.get('total_wp_earned') or 0
        has_wallet_data = deposit > 0 or balance > 0 or wp_earned > 0
        if has_wallet_data:
            total_spent = round(max(0, deposit - balance - wp_earned), 2)
            p['total_spent'] = total_spent
            p['pnl'] = round(wp_earned + balance - deposit, 2)
        else:
            manual_spent = p.get('spent') or 0
            p['total_spent'] = round(manual_spent, 2)
            p['pnl'] = round((p.get('earned') or 0) - manual_spent, 2)
    return jsonify(protocols)


@app.route('/api/protocols', methods=['POST'])
def add_protocol():
    data = request.json
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Название обязательно'}), 400
    try:
        db.add_protocol(
            name=name,
            spent=float(data.get('spent') or 0),
            earned=float(data.get('earned') or 0),
            status=data.get('status', 'фармлю'),
            note=data.get('note', ''),
            color=data.get('color', '#10b981')
        )
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/protocols/<int:pid>', methods=['PATCH'])
def patch_protocol(pid):
    data = request.json
    protocols = {p['id']: p for p in db.get_protocols()}
    p = protocols.get(pid)
    if not p:
        return jsonify({'error': 'Not found'}), 404
    db.update_protocol(
        pid=pid,
        spent=float(data.get('spent', p['spent']) or 0),
        earned=float(data.get('earned', p['earned']) or 0),
        status=data.get('status', p['status']),
        note=data.get('note', p['note'] or ''),
        color=data.get('color', p.get('color', '#10b981')),
        points=float(data.get('points', p.get('points') or 0) or 0),
        point_price=float(data.get('point_price', p.get('point_price') or 0) or 0),
    )
    pnl = round(float(data.get('earned', p['earned']) or 0) - float(data.get('spent', p['spent']) or 0), 2)
    return jsonify({'ok': True, 'pnl': pnl})


@app.route('/api/protocols/<int:pid>', methods=['PUT'])
def update_protocol(pid):
    data = request.json
    db.update_protocol(
        pid=pid,
        spent=float(data.get('spent') or 0),
        earned=float(data.get('earned') or 0),
        status=data.get('status', 'фармлю'),
        note=data.get('note', ''),
        color=data.get('color', '#10b981'),
        points=float(data.get('points') or 0),
        point_price=float(data.get('point_price') or 0),
    )
    return jsonify({'ok': True})


@app.route('/api/protocols/<int:pid>', methods=['DELETE'])
def delete_protocol(pid):
    db.delete_protocol(pid)
    return jsonify({'ok': True})


@app.route('/api/protocols/<path:name>/wallets', methods=['GET'])
def get_protocol_wallets_detail(name):
    return jsonify(db.get_protocol_wallets_detail(name))


@app.route('/api/protocols/<path:name>/import', methods=['POST'])
def import_protocol_wallets(name):
    rows = request.json.get('rows', [])
    add_points = bool(request.json.get('add_points', False))
    add_deposit = bool(request.json.get('add_deposit', False))
    week = request.json.get('week') or None
    result = db.import_protocol_wallet_data(name, rows, add_points=add_points, add_deposit=add_deposit, week=week)
    return jsonify({'ok': True, **result})


@app.route('/api/protocols/<path:name>/weeks', methods=['GET'])
def get_protocol_weeks(name):
    return jsonify(db.get_protocol_weeks(name))


@app.route('/api/protocols/<path:name>/weeks', methods=['POST'])
def add_protocol_week(name):
    data = request.json
    wid = db.add_protocol_week(name, data['week_start'], data['week_end'])
    return jsonify({'ok': True, 'id': wid})


@app.route('/api/protocols/<path:name>/weeks/<int:wid>', methods=['DELETE'])
def delete_protocol_week(name, wid):
    db.delete_protocol_week(wid)
    return jsonify({'ok': True})


@app.route('/api/protocols/<path:name>/clear-data', methods=['POST'])
def clear_protocol_data(name):
    db.clear_protocol_wallet_data(name)
    return jsonify({'ok': True})


@app.route('/api/protocols/<path:name>/weeks/<int:wid>', methods=['PATCH'])
def patch_protocol_week(name, wid):
    data = request.json
    week_start = data.get('week_start', '').strip()
    week_end = data.get('week_end', '').strip()
    if not week_start or not week_end:
        return jsonify({'error': 'Укажите даты'}), 400
    db.update_protocol_week(wid, week_start, week_end)
    return jsonify({'ok': True})


@app.route('/api/protocols/<path:name>/weeks/<week_start>/wallets', methods=['GET'])
def get_week_wallets(name, week_start):
    return jsonify(db.get_week_wallets(name, week_start))


@app.route('/api/wallet-protocols', methods=['PATCH'])
def update_wallet_protocol():
    data = request.json
    db.update_wallet_protocol(
        wallet_id=int(data['wallet_id']),
        protocol=data['protocol'],
        deposit=float(data['deposit']) if 'deposit' in data else None,
        wallet_balance=float(data['wallet_balance']) if 'wallet_balance' in data else None,
        wp_points=float(data['wp_points']) if 'wp_points' in data else None,
        wp_earned=float(data['wp_earned']) if 'wp_earned' in data else None,
    )
    return jsonify({'ok': True})


# --- Wallets ---

@app.route('/api/wallets', methods=['GET'])
def get_wallets():
    protocol = request.args.get('protocol')
    unassigned = request.args.get('unassigned') == '1'
    chain = request.args.get('chain') or None
    return jsonify(db.get_wallets(protocol=protocol or None, unassigned_only=unassigned, chain=chain))


@app.route('/api/wallets', methods=['POST'])
def add_wallet():
    data = request.json
    protocol = (data.get('protocol') or '').strip()
    address = (data.get('address') or '').strip()
    if not protocol or not address:
        return jsonify({'error': 'Протокол и адрес обязательны'}), 400
    db.bulk_add_wallets([address], protocols=[protocol], label=data.get('label', ''))
    return jsonify({'ok': True})



@app.route('/api/wallets/bulk', methods=['POST'])
def bulk_add_wallets():
    data = request.json
    raw = data.get('addresses', '')
    addresses = [a.strip() for a in (raw if isinstance(raw, list) else raw.split('\n')) if a.strip()]
    if not addresses:
        return jsonify({'error': 'Нет адресов'}), 400
    protocols = [p for p in (data.get('protocols') or []) if p]
    count = db.bulk_add_wallets(addresses, protocols=protocols or None, label=data.get('label', ''))
    return jsonify({'ok': True, 'count': count})


@app.route('/api/wallets/assign', methods=['PUT'])
def assign_wallets():
    data = request.json
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'error': 'Нет кошельков'}), 400

    if 'protocols' in data:
        # Replace all protocols for each wallet (single-wallet assign modal)
        for wid in ids:
            db.set_wallet_protocols(wid, data['protocols'])
    else:
        # Add one protocol to multiple wallets (bulk bar action)
        db.add_protocol_to_wallets(ids, data.get('protocol'))

    return jsonify({'ok': True})


@app.route('/api/wallets/<int:wid>', methods=['PATCH'])
def update_wallet(wid):
    data = request.json or {}
    if 'label' in data:
        db.update_wallet_label(wid, data.get('label', ''))
    if 'proxy' in data:
        db.update_wallet_proxy(wid, data.get('proxy', ''))
    if 'chain' in data:
        import sqlite3 as _sq
        with db._conn() as c:
            c.execute('UPDATE wallets SET chain=? WHERE id=?', (data.get('chain') or None, wid))
    return jsonify({'ok': True})



@app.route('/api/wallets/import-balances', methods=['POST'])
def import_balances():
    import openpyxl, io
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'Нет файла'}), 400
    wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    rows = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue
            rows.append({
                'address_short': str(row[1]),
                'volume':  row[2],
                'burn':    row[3],
                'points':  row[4],
                'p_price': row[5],
                'balance': row[6],
            })
    matched = db.import_balances(rows)
    return jsonify({'ok': True, 'matched': matched, 'total': len(rows)})


@app.route('/api/wallets/bulk-delete', methods=['POST'])
def bulk_delete_wallets():
    ids = request.json.get('ids', [])
    for wid in ids:
        db.delete_wallet(wid)
    return jsonify({'ok': True, 'count': len(ids)})


@app.route('/api/wallets/<int:wid>', methods=['DELETE'])
def delete_wallet(wid):
    db.delete_wallet(wid)
    return jsonify({'ok': True})


# --- Reminders ---

@app.route('/api/reminders', methods=['GET'])
def get_reminders():
    include_done = request.args.get('include_done') == '1'
    return jsonify(db.get_reminders(include_done=include_done))


@app.route('/api/reminders', methods=['POST'])
def add_reminder():
    data = request.json
    protocol = (data.get('protocol') or '').strip()
    remind_at = (data.get('remind_at') or '').strip()
    if not protocol or not remind_at:
        return jsonify({'error': 'Протокол и дата обязательны'}), 400
    db.add_reminder(protocol, remind_at)
    return jsonify({'ok': True})


@app.route('/api/reminders/<int:rid>', methods=['DELETE'])
def delete_reminder(rid):
    db.delete_reminder(rid)
    return jsonify({'ok': True})


@app.route('/api/reminders/<int:rid>/done', methods=['POST'])
def mark_done(rid):
    db.mark_done(rid)
    return jsonify({'ok': True})


@app.route('/api/due-reminders')
def due_reminders():
    return jsonify(db.get_due_reminders())


# --- Extended ---

EXTENDED_API = 'https://api.starknet.extended.exchange'


def _fetch_extended_account(api_key: str) -> dict:
    headers = {'X-Api-Key': api_key}

    bal_resp = _get_json(f'{EXTENDED_API}/api/v1/user/balance', headers=headers, timeout=10)
    pos_resp = _get_json(f'{EXTENDED_API}/api/v1/user/positions', headers=headers, timeout=10)

    balance = bal_resp.get('data', {}) if isinstance(bal_resp.get('data'), dict) else {}
    positions_raw = pos_resp.get('data', []) if isinstance(pos_resp.get('data'), list) else []

    equity = float(balance.get('equity', 0) or 0)
    available = float(balance.get('availableForTrade', 0) or 0)
    unrealised_pnl = float(balance.get('unrealisedPnl', 0) or 0)

    positions = []
    total_margin = 0.0
    total_notional = 0.0
    for p in positions_raw:
        size = float(p.get('size', 0) or 0)
        if size == 0:
            continue
        entry = float(p.get('openPrice', 0) or 0)
        mark = float(p.get('markPrice', 0) or 0)
        liq = p.get('liquidationPrice')
        upnl = float(p.get('unrealisedPnl', 0) or 0)
        margin = float(p.get('margin', 0) or 0)
        lev = p.get('leverage', {})
        lev_val = float(lev.get('value', 0) or 0) if isinstance(lev, dict) else 0
        direction = 'long' if (p.get('side', '') or '').upper() == 'LONG' else 'short'
        total_margin += margin
        total_notional += abs(size) * (mark or entry)
        dist_liq = None
        if liq and mark:
            liq_f = float(liq)
            if direction == 'long':
                dist_liq = round((mark - liq_f) / mark * 100, 2) if mark else None
            else:
                dist_liq = round((liq_f - mark) / mark * 100, 2) if mark else None
        positions.append({
            'market': p.get('market', ''),
            'direction': direction,
            'size': round(size, 6),
            'entry_price': round(entry, 6),
            'mark_price': round(mark, 6),
            'liq_price': round(float(liq), 6) if liq else None,
            'unrealised_pnl': round(upnl, 4),
            'margin': round(margin, 4),
            'leverage': round(lev_val, 1),
            'dist_liq': dist_liq,
        })

    return {
        'positions': positions,
        'equity': round(equity, 2),
        'available': round(available, 2),
        'margin_used': round(total_margin, 2),
        'unrealised_pnl': round(unrealised_pnl, 4),
        'account_leverage': round(total_notional / equity, 2) if equity > 0 and total_notional > 0 else None,
        **_account_risk_metrics(equity, total_margin, available),
    }


@app.route('/api/extended/accounts', methods=['GET'])
def get_extended_accounts():
    return jsonify(db.get_extended_accounts())


@app.route('/api/extended/accounts', methods=['POST'])
def add_extended_account():
    data = request.json
    label = (data.get('label') or '').strip()
    api_key = (data.get('api_key') or '').strip()
    if not label or not api_key:
        return jsonify({'error': 'label and api_key required'}), 400
    db.add_extended_account(label, api_key)
    return jsonify({'ok': True})


@app.route('/api/extended/accounts/<int:aid>', methods=['DELETE'])
def delete_extended_account(aid):
    db.delete_extended_account(aid)
    return jsonify({'ok': True})


@app.route('/api/extended/accounts/<int:aid>', methods=['PATCH'])
def patch_extended_account(aid):
    data = request.json or {}
    if 'group_name' in data:
        db.update_extended_account_group(aid, data.get('group_name'))
    return jsonify({'ok': True})


@app.route('/api/extended/all-positions')
def extended_all_positions():
    accounts = db.get_extended_accounts()
    if not accounts:
        return jsonify({'accounts': []})
    result = []
    for acc in accounts:
        try:
            data = _fetch_extended_account(acc['api_key'])
            data['id'] = acc['id']
            data['label'] = acc['label']
            data['group_name'] = acc.get('group_name')
            result.append(data)
        except Exception as e:
            result.append({'id': acc['id'], 'label': acc['label'],
                           'error': str(e), 'positions': []})
    return jsonify({'accounts': result})


# --- Pacifica ---

PACIFICA_API = 'https://api.pacifica.fi/api/v1'
_pacifica_prices: dict = {}
_pacifica_prices_ts: float = 0

def _pacifica_get_prices() -> dict:
    """Returns {symbol: mark_price}, cached for 10 seconds."""
    global _pacifica_prices, _pacifica_prices_ts
    now = time.time()
    if _pacifica_prices and (now - _pacifica_prices_ts) < 10:
        return _pacifica_prices
    resp = _get_json(f'{PACIFICA_API}/info/prices', timeout=10)
    prices = {}
    for item in (resp.get('data') or []):
        sym = item.get('symbol', '')
        try:
            prices[sym] = float(item.get('mark', 0) or 0)
        except Exception:
            prices[sym] = 0
    _pacifica_prices = prices
    _pacifica_prices_ts = now
    return prices


def _fetch_pacifica_account(address: str) -> dict:
    prices = _pacifica_get_prices()

    pos_resp = _get_json(f'{PACIFICA_API}/positions', params={'account': address}, timeout=10)
    acc_resp = _get_json(f'{PACIFICA_API}/account', params={'account': address}, timeout=10)

    positions_raw = pos_resp.get('data') or []
    account_data = acc_resp.get('data') or {}

    margin_used = float(account_data.get('total_margin_used', 0) or 0)
    equity = float(account_data.get('account_equity', 0) or 0)
    maint_margin = float(account_data.get('cross_mmr') or account_data.get('maintenance_margin') or 0)

    positions = []
    for p in positions_raw:
        sym = p.get('symbol', '')
        side_raw = p.get('side', '')
        side = 'Long' if side_raw == 'bid' else 'Short'
        amount = float(p.get('amount', 0) or 0)
        entry = float(p.get('entry_price', 0) or 0)
        mark = prices.get(sym, 0)
        funding = float(p.get('funding', 0) or 0)
        isolated = p.get('isolated', False)
        # Use API liq_price if provided
        api_liq = p.get('liq_price') or p.get('liquidation_price')

        if mark and amount:
            raw_pnl = (mark - entry) * amount if side == 'Long' else (entry - mark) * amount
            pnl = raw_pnl + funding
        else:
            pnl = funding

        mark_val = amount * (mark if mark else entry)
        positions.append({
            'market': sym,
            'side': side,
            'size': amount,
            'entry_price': entry,
            'mark_price': mark,
            'unrealised_pnl': pnl,
            'funding': funding,
            'isolated': isolated,
            'mark_value': mark_val,   # used for margin/liq allocation below
            'api_liq': float(api_liq) if api_liq else None,
        })

    # Per-position margin, leverage, and liq price
    # Margin is allocated proportional to mark value (verified against Pacifica UI)
    # Liq price uses cross-margin formula: mark ± (equity - pos_maint_margin) / size
    # where pos_maint_margin is proportionally allocated from account maintenance_margin
    total_mark_value = sum(p['mark_value'] for p in positions)
    for p in positions:
        mv = p['mark_value']
        if margin_used > 0 and total_mark_value > 0 and mv > 0:
            weight = mv / total_mark_value
            pos_margin = margin_used * weight
            leverage = round(mv / pos_margin, 1)
            # Use API liq_price if available, otherwise calculate cross-margin liq
            if p['api_liq'] is not None:
                liq = p['api_liq']
            elif p['size'] > 0:
                pos_maint = maint_margin * weight if maint_margin else 0
                effective_equity = equity - pos_maint
                if p['side'] == 'Long':
                    liq = round((p['mark_price'] or p['entry_price']) - effective_equity / p['size'], 4)
                else:
                    liq = round((p['mark_price'] or p['entry_price']) + effective_equity / p['size'], 4)
            else:
                liq = None
        else:
            pos_margin = None
            leverage = None
            liq = p['api_liq']
        p['margin'] = round(pos_margin, 4) if pos_margin is not None else None
        p['leverage'] = leverage
        p['liq_price'] = liq
        del p['mark_value']
        del p['api_liq']

    account_leverage = round(total_mark_value / equity, 2) if equity > 0 and total_mark_value > 0 else None
    margin_ratio = round(maint_margin / equity * 100, 2) if equity > 0 and maint_margin > 0 else None
    risk = _account_risk_metrics(equity, margin_used, float(account_data.get('available_to_spend', 0) or 0))

    return {
        'positions': positions,
        'balance': float(account_data.get('balance', 0) or 0),
        'equity': equity,
        'available': float(account_data.get('available_to_spend', 0) or 0),
        'margin_used': margin_used,
        'maint_margin': maint_margin,
        'unrealised_pnl': sum(p['unrealised_pnl'] for p in positions),
        'account_leverage': account_leverage,
        'margin_ratio': margin_ratio,
        **risk,
    }


@app.route('/api/pacifica/accounts', methods=['GET'])
def get_pacifica_accounts():
    return jsonify(db.get_pacifica_accounts())


@app.route('/api/pacifica/accounts', methods=['POST'])
def add_pacifica_account():
    data = request.json
    label = (data.get('label') or '').strip()
    address = (data.get('address') or '').strip()
    if not label or not address:
        return jsonify({'error': 'label and address required'}), 400
    db.add_pacifica_account(label, address)
    return jsonify({'ok': True})


@app.route('/api/pacifica/accounts/<int:aid>', methods=['DELETE'])
def delete_pacifica_account(aid):
    db.delete_pacifica_account(aid)
    return jsonify({'ok': True})


@app.route('/api/pacifica/accounts/<int:aid>', methods=['PATCH'])
def patch_pacifica_account(aid):
    data = request.json or {}
    if 'group_name' in data:
        db.update_pacifica_account_group(aid, data['group_name'])
    return jsonify({'ok': True})


@app.route('/api/pacifica/all-positions')
def pacifica_all_positions():
    accounts = db.get_pacifica_accounts()
    if not accounts:
        return jsonify({'accounts': []})
    result = []
    for acc in accounts:
        try:
            data = _fetch_pacifica_account(acc['address'])
            data['id'] = acc['id']
            data['label'] = acc['label']
            data['group_name'] = acc.get('group_name')
            result.append(data)
        except Exception as e:
            result.append({'id': acc['id'], 'label': acc['label'],
                           'error': str(e), 'positions': []})
    return jsonify({'accounts': result})


# --- Ethereal ---

ETHEREAL_API = 'https://api.ethereal.trade'


def _ethereal_get_products() -> dict:
    """Returns {product_id: symbol}, cached for 5 minutes."""
    global _ethereal_products, _ethereal_products_ts
    now = time.time()
    if _ethereal_products and (now - _ethereal_products_ts) < 300:
        return _ethereal_products
    try:
        resp = _get_json(f'{ETHEREAL_API}/v1/product', params={'limit': 100}, timeout=10)
        mapping = {}
        products = resp if isinstance(resp, list) else resp.get('data', [])
        for p in products:
            pid = p.get('id') or p.get('productId')
            sym = p.get('displayTicker') or p.get('ticker') or p.get('symbol') or p.get('name') or str(pid)
            if pid:
                mapping[str(pid)] = sym
        _ethereal_products = mapping
        _ethereal_products_ts = now
    except Exception:
        pass
    return _ethereal_products


def _ethereal_data_list(resp) -> list:
    if isinstance(resp, list):
        return resp
    data = resp.get('data') if isinstance(resp, dict) else None
    return data if isinstance(data, list) else []


def _ethereal_float(value, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _ethereal_direction(side) -> str:
    if side in (0, '0'):
        return 'long'
    if side in (1, '1'):
        return 'short'
    side_s = str(side or '').upper()
    return 'long' if side_s in ('BUY', 'LONG') else 'short'


def _ethereal_get_market_prices(product_ids: list[str]) -> dict:
    ids = sorted({pid for pid in product_ids if pid})
    if not ids:
        return {}
    try:
        resp = _get_json(
            f'{ETHEREAL_API}/v1/product/market-price',
            params=[('productIds[]', pid) for pid in ids],
            timeout=10,
        )
    except Exception:
        return {}

    prices = {}
    for item in _ethereal_data_list(resp):
        pid = str(item.get('productId') or item.get('product_id') or '')
        price = (
            item.get('oraclePrice')
            or item.get('markPrice')
            or item.get('bestBidPrice')
            or item.get('bestAskPrice')
        )
        if pid and price is not None:
            prices[pid] = _ethereal_float(price)
    return prices


def _fetch_ethereal_account(address: str) -> dict:
    products = _ethereal_get_products()

    # Step 1: resolve subaccounts for this sender
    sub_resp = _get_json(f'{ETHEREAL_API}/v1/subaccount', params={'sender': address}, timeout=10)
    subaccounts = _ethereal_data_list(sub_resp)
    if not subaccounts:
        return {'positions': [], 'equity': 0.0, 'available': 0.0, 'margin_used': 0.0, 'unrealised_pnl': 0.0}

    all_positions = []
    total_equity = 0.0
    total_available = 0.0
    total_margin = 0.0
    total_pnl = 0.0

    for sub in subaccounts:
        sub_id = sub.get('id') or sub.get('subaccountId')
        if not sub_id:
            continue

        sub_used = 0.0

        # Positions
        try:
            pos_resp = _get_json(
                f'{ETHEREAL_API}/v1/position',
                params={'subaccountId': sub_id, 'open': 'true'},
                timeout=10,
            )
            raw_positions = _ethereal_data_list(pos_resp)
        except Exception:
            raw_positions = []

        # Balance
        try:
            bal_resp = _get_json(
                f'{ETHEREAL_API}/v1/subaccount/balance',
                params={'subaccountId': sub_id},
                timeout=10,
            )
            balances = _ethereal_data_list(bal_resp)
            usd_balance = next((b for b in balances if b.get('tokenName') == 'USD'), balances[0] if balances else {})
            equity    = _ethereal_float(usd_balance.get('amount'))
            available = _ethereal_float(usd_balance.get('available'))
            used      = _ethereal_float(usd_balance.get('totalUsed'))
            sub_used  = used
            total_equity    += equity
            total_available += available
            total_margin    += used
        except Exception:
            pass

        product_ids = [str(p.get('productId') or p.get('product_id') or '') for p in raw_positions]
        prices = _ethereal_get_market_prices(product_ids)
        position_costs = [
            abs(_ethereal_float(p.get('cost') or p.get('positionValue') or p.get('position_value')))
            for p in raw_positions
        ]
        cost_total = sum(position_costs)

        for p in raw_positions:
            product_id = str(p.get('productId') or p.get('product_id') or '')
            symbol = products.get(product_id, product_id or '?')
            direction = _ethereal_direction(p.get('side'))

            size = abs(_ethereal_float(p.get('size')))
            cost = abs(_ethereal_float(p.get('cost') or p.get('positionValue') or p.get('position_value')))
            entry = _ethereal_float(p.get('entryPrice') or p.get('entry_price'))
            if not entry and size:
                entry = cost / size

            mark = (
                prices.get(product_id)
                or _ethereal_float(p.get('markPrice') or p.get('mark_price') or p.get('currentPrice') or p.get('current_price'))
            )
            liq = _ethereal_float(p.get('liquidationPrice') or p.get('liquidation_price'))
            pnl = _ethereal_float(p.get('unrealizedPnl') or p.get('unrealized_pnl'))
            total_pnl += pnl

            dist_liq = None
            if mark and liq:
                dist_liq = round(
                    (mark - liq) / mark * 100 if direction == 'long' else (liq - mark) / mark * 100,
                    2,
                )

            margin = _ethereal_float(p.get('marginUsage') or p.get('margin_usage'))
            if not margin and cost_total:
                margin = sub_used * (cost / cost_total)
            leverage = cost / margin if margin else None

            all_positions.append({
                'symbol':          symbol,
                'direction':       direction,
                'size':            size,
                'entry_price':     entry or None,
                'current_price':   mark,
                'mark_price':      mark,
                'liq_price':       liq or None,
                'dist_liq':        dist_liq,
                'unrealized_pnl':  pnl,
                'margin':          round(margin, 6) if margin else 0.0,
                'leverage':        round(leverage, 4) if leverage else None,
            })

    return {
        'positions':     all_positions,
        'equity':        round(total_equity, 2),
        'available':     round(total_available, 2),
        'margin_used':   round(total_margin, 2),
        'unrealised_pnl': round(total_pnl, 4),
        'account_leverage': (
            round(sum(abs(p['size']) * (p['current_price'] or p['entry_price'] or 0) for p in all_positions) / total_equity, 2)
            if total_equity > 0 and all_positions else None
        ),
        **_account_risk_metrics(total_equity, total_margin, total_available),
    }


@app.route('/api/ethereal/accounts', methods=['GET'])
def get_ethereal_accounts():
    return jsonify(db.get_ethereal_accounts())


@app.route('/api/ethereal/accounts', methods=['POST'])
def add_ethereal_account():
    data = request.json
    label   = (data.get('label')   or '').strip()
    address = (data.get('address') or '').strip()
    if not label or not address:
        return jsonify({'error': 'label and address required'}), 400
    db.add_ethereal_account(label, address)
    return jsonify({'ok': True})


@app.route('/api/ethereal/accounts/<int:aid>', methods=['DELETE'])
def delete_ethereal_account(aid):
    db.delete_ethereal_account(aid)
    return jsonify({'ok': True})


@app.route('/api/ethereal/accounts/<int:aid>', methods=['PATCH'])
def patch_ethereal_account(aid):
    data = request.json or {}
    if 'group_name' in data:
        db.update_ethereal_account_group(aid, data.get('group_name'))
    return jsonify({'ok': True})


@app.route('/api/ethereal/all-positions')
def ethereal_all_positions():
    accounts = db.get_ethereal_accounts()
    if not accounts:
        return jsonify({'accounts': []})
    result = []
    for acc in accounts:
        try:
            data = _fetch_ethereal_account(acc['address'])
            data['id']         = acc['id']
            data['label']      = acc['label']
            data['address']    = acc['address']
            data['group_name'] = acc.get('group_name')
            result.append(data)
        except Exception as e:
            result.append({'id': acc['id'], 'label': acc['label'],
                           'address': acc['address'], 'error': str(e), 'positions': []})
    return jsonify({'accounts': result})


# --- Nado ---

NADO_API = 'https://gateway.prod.nado.xyz/v1'
_nado_symbols: dict = {}
_nado_symbols_ts: float = 0

def _nado_get_symbols() -> dict:
    """Returns {product_id: symbol_name}, cached for 5 minutes."""
    global _nado_symbols, _nado_symbols_ts
    now = time.time()
    if _nado_symbols and (now - _nado_symbols_ts) < 300:
        return _nado_symbols
    try:
        resp = _get_json(f'{NADO_API}/query', params={'type': 'symbols'}, timeout=10)
        mapping = {}
        for sym, info in resp.get('data', {}).get('symbols', {}).items():
            mapping[info['product_id']] = sym
        _nado_symbols = mapping
        _nado_symbols_ts = now
    except Exception:
        pass
    return _nado_symbols

def _addr_to_nado_subaccount(address: str) -> str:
    """Convert EVM address to Nado default subaccount (bytes32 hex).
    Format: address (20 bytes) + 'default' in ASCII (7 bytes) + zero padding (5 bytes)
    """
    addr = address.lower().replace('0x', '').zfill(40)
    default_hex = 'default'.encode().hex()  # 64656661756c74 = 7 bytes
    padding = '00' * 5                       # 5 zero bytes
    return '0x' + addr + default_hex + padding


def _fetch_nado_account(address: str) -> dict:
    subaccount = _addr_to_nado_subaccount(address)
    symbols = _nado_get_symbols()

    # Fetch positions and account info in parallel would be ideal, but sequential is fine
    pos_resp = _get_json(
        f'{NADO_API}/query',
        params={'type': 'isolated_positions', 'subaccount': subaccount},
        timeout=10
    )

    # Fetch overall account info (value + available margin)
    account_value = None
    available_margin = None
    try:
        info_resp = _get_json(
            f'{NADO_API}/query',
            params={'type': 'subaccount_info', 'subaccount': subaccount},
            timeout=10
        )
        healths = info_resp.get('data', {}).get('healths', [])
        if healths:
            account_value = round(int(healths[0].get('assets', 0)) / 1e18, 2)
            available_margin = round(int(healths[0].get('health', 0)) / 1e18, 2)
    except Exception:
        pass

    positions = []
    if pos_resp.get('status') == 'success':
        for p in pos_resp.get('data', {}).get('isolated_positions', []):
            base = p.get('base_balance', {}).get('balance', {})
            quote = p.get('quote_balance', {}).get('balance', {})
            base_product = p.get('base_product', {})
            product_id = p.get('base_balance', {}).get('product_id')
            amount_raw = int(base.get('amount', 0))
            v_quote_raw = int(base.get('v_quote_balance', 0))
            quote_raw = int(quote.get('amount', 0))
            if amount_raw == 0:
                continue
            size = amount_raw / 1e18
            direction = 'long' if size > 0 else 'short'
            size_abs = abs(size)
            entry_price = abs(v_quote_raw / 1e18) / size_abs if size_abs else 0

            # Current oracle price
            oracle_raw = base_product.get('oracle_price_x18', 0)
            current_price = int(oracle_raw) / 1e18 if oracle_raw else None

            # Unrealized PnL
            unrealized_pnl = None
            if current_price and entry_price:
                sign = 1 if direction == 'long' else -1
                unrealized_pnl = round((current_price - entry_price) * size_abs * sign, 4)

            # Health % and liq price from healths[1] (maintenance)
            health_pct = None
            liq_price = None
            pos_healths = p.get('healths', [])
            if len(pos_healths) >= 2:
                maint = pos_healths[1]
                h_assets = int(maint.get('assets', 0))
                h_health = int(maint.get('health', 0))
                if h_assets > 0:
                    health_pct = round(h_health / h_assets * 100, 2)
                # Liq price: price at which maintenance health reaches 0
                # dHealth/dP = size * W_maint (long) or -size * W_maint (short)
                # P_liq = current ∓ health / (size * W_maint)
                risk = base_product.get('risk', {})
                if current_price and size_abs and h_health >= 0:
                    health_usd = h_health / 1e18
                    if direction == 'long':
                        w = int(risk.get('long_weight_maintenance_x18', 0)) / 1e18
                        if w:
                            liq_price = round(current_price - health_usd / (size_abs * w), 4)
                    else:
                        w = int(risk.get('short_weight_maintenance_x18', 0)) / 1e18
                        if w:
                            liq_price = round(current_price + health_usd / (size_abs * w), 4)

            margin = quote_raw / 1e18
            leverage = round((size_abs * entry_price) / margin, 1) if margin > 0 and entry_price > 0 else None
            symbol = symbols.get(product_id, f'#{product_id}')
            positions.append({
                'product_id': product_id,
                'symbol': symbol,
                'direction': direction,
                'size': round(size_abs, 6),
                'entry_price': round(entry_price, 6),
                'current_price': round(current_price, 6) if current_price else None,
                'unrealized_pnl': unrealized_pnl,
                'health_pct': health_pct,
                'liq_price': liq_price,
                'margin': round(margin, 4),
                'leverage': leverage,
            })
    # Total equity = available margin + sum of (margin + unrealized_pnl) per isolated position
    # This matches Nado's "Total Equity" display
    if available_margin is not None:
        pos_equity = sum((p['margin'] + (p['unrealized_pnl'] or 0)) for p in positions)
        total_equity = round(available_margin + pos_equity, 2)
    else:
        total_equity = None
    margin_used = round(sum(p['margin'] for p in positions), 4)
    account_leverage = None
    if total_equity and total_equity > 0:
        notional = sum(abs(p['size']) * (p['current_price'] or p['entry_price'] or 0) for p in positions)
        account_leverage = round(notional / total_equity, 2) if notional > 0 else None

    return {
        'positions': positions,
        'subaccount': subaccount,
        'account_value': total_equity,
        'available_margin': available_margin,
        'margin_used': margin_used,
        'account_leverage': account_leverage,
        **_account_risk_metrics(total_equity, margin_used, available_margin),
    }


@app.route('/api/nado/accounts', methods=['GET'])
def get_nado_accounts():
    return jsonify(db.get_nado_accounts())


@app.route('/api/nado/accounts', methods=['POST'])
def add_nado_account():
    data = request.json
    label = (data.get('label') or '').strip()
    address = (data.get('address') or '').strip()
    if not label or not address:
        return jsonify({'error': 'label and address required'}), 400
    db.add_nado_account(label, address)
    return jsonify({'ok': True})


@app.route('/api/nado/accounts/<int:aid>', methods=['DELETE'])
def delete_nado_account(aid):
    db.delete_nado_account(aid)
    return jsonify({'ok': True})


@app.route('/api/nado/accounts/<int:aid>', methods=['PATCH'])
def patch_nado_account(aid):
    data = request.json or {}
    if 'group_name' in data:
        db.update_nado_account_group(aid, data.get('group_name'))
    return jsonify({'ok': True})


@app.route('/api/nado/all-positions')
def nado_all_positions():
    accounts = db.get_nado_accounts()
    if not accounts:
        return jsonify({'accounts': []})
    result = []
    for acc in accounts:
        try:
            data = _fetch_nado_account(acc['address'])
            data['id'] = acc['id']
            data['label'] = acc['label']
            data['address'] = acc['address']
            data['group_name'] = acc.get('group_name')
            result.append(data)
        except Exception as e:
            result.append({'id': acc['id'], 'label': acc['label'],
                           'address': acc['address'], 'error': str(e),
                           'positions': []})
    return jsonify({'accounts': result})


# --- Hyperliquid ---

HL_API = 'https://api.hyperliquid.xyz/info'

_mids_cache: dict = {}
_mids_cache_ts: float = 0


def _fetch_hl_account(address, mids):
    state = _post_json(HL_API, json={'type': 'clearinghouseState', 'user': address}, timeout=10)
    positions = []
    for ap in state.get('assetPositions', []):
        p = ap['position']
        szi = float(p['szi'])
        if szi == 0:
            continue
        coin = p['coin']
        direction = 'long' if szi > 0 else 'short'
        cur = float(mids[coin]) if coin in mids else None
        liq = float(p['liquidationPx']) if p.get('liquidationPx') else None
        dist_liq = None
        if cur and liq:
            dist_liq = round(((cur - liq) / cur * 100) if direction == 'long' else ((liq - cur) / cur * 100), 2)
        positions.append({
            'coin': coin, 'direction': direction, 'size': abs(szi),
            'entry_price': float(p['entryPx']), 'current_price': cur,
            'liq_price': liq, 'dist_liq': dist_liq,
            'margin_used': float(p['marginUsed']),
            'position_value': float(p['positionValue']),
            'unrealized_pnl': float(p['unrealizedPnl']),
            'leverage': p['leverage']['value'],
            'leverage_type': p['leverage']['type'],
            'roe': round(float(p.get('returnOnEquity', 0)) * 100, 2),
        })
    ms = state.get('marginSummary', {})
    account_value = float(ms.get('accountValue', 0))
    total_margin = float(ms.get('totalMarginUsed', 0))
    withdrawable = float(state.get('withdrawable', 0))
    return {
        'positions': positions,
        'summary': {
            'account_value': account_value,
            'total_margin':  total_margin,
            'total_pnl':     float(ms.get('totalUnrealizedPnl', 0)),
            'withdrawable':  withdrawable,
            'account_leverage': (
                round(sum(p['position_value'] for p in positions) / account_value, 2)
                if account_value > 0 and positions else None
            ),
            **_account_risk_metrics(account_value, total_margin, withdrawable),
        }
    }


@app.route('/api/hl/accounts', methods=['GET'])
def get_hl_accounts():
    return jsonify(db.get_hl_accounts())


@app.route('/api/hl/accounts', methods=['POST'])
def add_hl_account():
    data = request.json
    label = (data.get('label') or '').strip()
    address = (data.get('address') or '').strip()
    if not label or not address:
        return jsonify({'error': 'label and address required'}), 400
    db.add_hl_account(label, address)
    return jsonify({'ok': True})


@app.route('/api/hl/accounts/<int:aid>', methods=['DELETE'])
def delete_hl_account(aid):
    db.delete_hl_account(aid)
    return jsonify({'ok': True})


@app.route('/api/hl/accounts/<int:aid>', methods=['PATCH'])
def patch_hl_account(aid):
    data = request.json or {}
    if 'group_name' in data:
        db.update_hl_account_group(aid, data.get('group_name'))
    return jsonify({'ok': True})



@app.route('/api/hl/all-positions')
def hl_all_positions():
    global _mids_cache, _mids_cache_ts
    accounts = db.get_hl_accounts()
    if not accounts:
        return jsonify({'accounts': [], 'mids_ts': None})
    try:
        now = time.time()
        if not _mids_cache or (now - _mids_cache_ts) > _PRICE_TTL:
            _mids_cache = _post_json(HL_API, json={'type': 'allMids'}, timeout=10)
            _mids_cache_ts = now
        mids = _mids_cache

        result = []
        for acc in accounts:
            try:
                data = _fetch_hl_account(acc['address'], mids)
                data['id'] = acc['id']
                data['label'] = acc['label']
                data['address'] = acc['address']
                result.append(data)
            except Exception as e:
                result.append({'id': acc['id'], 'label': acc['label'],
                               'address': acc['address'], 'error': str(e),
                               'positions': [], 'summary': {}})
        return jsonify({'accounts': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 502


# --- Calculator (manual positions) ---

@app.route('/api/positions', methods=['GET'])
def get_positions():
    return jsonify(db.get_positions())


@app.route('/api/positions', methods=['POST'])
def add_position():
    data = request.json
    try:
        gid = data.get('group_id')
        db.add_position(
            symbol=data['symbol'],
            direction=data['direction'],
            collateral=float(data['collateral']),
            leverage=float(data['leverage']),
            entry_price=float(data['entry_price']),
            note=data.get('note', ''),
            group_id=int(gid) if gid else None,
        )
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/positions/<int:pid>', methods=['PATCH'])
def update_position(pid):
    data = request.json
    try:
        gid = data.get('group_id')
        db.update_position(
            pid=pid,
            symbol=data['symbol'],
            direction=data['direction'],
            collateral=float(data['collateral']),
            leverage=float(data['leverage']),
            entry_price=float(data['entry_price']),
            note=data.get('note', ''),
            group_id=int(gid) if gid else None,
        )
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/positions/<int:pid>', methods=['DELETE'])
def delete_position(pid):
    db.delete_position(pid)
    return jsonify({'ok': True})


# --- Position groups ---

@app.route('/api/position-groups', methods=['GET'])
def get_position_groups():
    return jsonify(db.get_position_groups())


@app.route('/api/position-groups', methods=['POST'])
def add_position_group():
    data = request.json
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Нужно имя'}), 400
    gid = db.add_position_group(name)
    return jsonify({'id': gid, 'ok': True})


@app.route('/api/position-groups/<int:gid>', methods=['PATCH'])
def rename_position_group(gid):
    data = request.json
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Нужно имя'}), 400
    db.rename_position_group(gid, name)
    return jsonify({'ok': True})


@app.route('/api/position-groups/<int:gid>', methods=['DELETE'])
def delete_position_group(gid):
    db.delete_position_group(gid)
    return jsonify({'ok': True})


@app.route('/api/hl/mids')
def hl_mids():
    """Return all current mid prices from Hyperliquid."""
    global _mids_cache, _mids_cache_ts
    try:
        now = time.time()
        if not _mids_cache or (now - _mids_cache_ts) > _PRICE_TTL:
            _mids_cache = _post_json(HL_API, json={'type': 'allMids'}, timeout=10)
            _mids_cache_ts = now
        return jsonify(_mids_cache)
    except Exception as e:
        return jsonify({'error': str(e)}), 502


@app.route('/api/hl/search')
def hl_search():
    """Search HL tokens by symbol prefix."""
    global _mids_cache, _mids_cache_ts
    q = request.args.get('q', '').strip().upper()
    try:
        now = time.time()
        if not _mids_cache or (now - _mids_cache_ts) > _PRICE_TTL:
            _mids_cache = _post_json(HL_API, json={'type': 'allMids'}, timeout=10)
            _mids_cache_ts = now
        results = [
            {'symbol': k, 'price': float(v)}
            for k, v in _mids_cache.items()
            if q in k.upper()
        ]
        results.sort(key=lambda x: (not x['symbol'].upper().startswith(q), x['symbol']))
        return jsonify(results[:10])
    except Exception as e:
        return jsonify({'error': str(e)}), 502


# --- Export ---

@app.route('/api/proxies', methods=['GET'])
def get_proxies():
    db.sync_proxies()
    return jsonify(db.get_proxies())


@app.route('/api/proxies/bulk', methods=['POST'])
def bulk_add_proxies():
    lines = (request.json.get('proxies') or [])
    added = db.bulk_add_proxies(lines)
    return jsonify({'ok': True, 'added': added})


@app.route('/api/proxies/<int:pid>', methods=['PATCH'])
def update_proxy(pid):
    data = request.json or {}
    if 'label' in data:
        db.update_proxy_label(pid, data['label'])
    if 'wallet_id' in data:
        wid = data['wallet_id']
        if wid:
            db.assign_proxy(pid, wid)
        else:
            db.unassign_proxy(pid)
    return jsonify({'ok': True})


@app.route('/api/proxies/<int:pid>', methods=['DELETE'])
def delete_proxy(pid):
    db.delete_proxy(pid)
    return jsonify({'ok': True})


@app.route('/api/open-folder', methods=['POST'])
def open_folder():
    import subprocess
    path = request.json.get('path', '')
    if path and os.path.exists(path):
        subprocess.Popen(['explorer', '/select,', os.path.normpath(path)])
    return jsonify({'ok': True})


@app.route('/api/export')
def export():
    wb = openpyxl.Workbook()

    dark = PatternFill('solid', fgColor='1a1a2e')
    light = PatternFill('solid', fgColor='F5F5F3')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin', color='2d3748'),
        right=Side(style='thin', color='2d3748'),
        top=Side(style='thin', color='2d3748'),
        bottom=Side(style='thin', color='2d3748'),
    )

    def style_header(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = dark
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin
        ws.row_dimensions[1].height = 28

    def style_rows(ws, start=2):
        for i, row in enumerate(ws.iter_rows(min_row=start)):
            if i % 2 == 0:
                for cell in row:
                    cell.fill = light
                    cell.border = thin

    # Sheet 1: Protocols
    ws1 = wb.active
    ws1.title = 'Протоколы'
    style_header(ws1, ['Протокол', 'Потрачено', 'Заработано', 'P&L', 'Кошельки', 'Статус', 'Заметка'])

    protocols = db.get_protocols()
    wallet_counts = db.get_wallet_counts()
    for p in protocols:
        pnl = round((p['earned'] or 0) - (p['spent'] or 0), 2)
        ws1.append([p['name'], p['spent'], p['earned'], pnl,
                    wallet_counts.get(p['name'], 0), p['status'], p['note'] or ''])

    style_rows(ws1)
    for i, row in enumerate(ws1.iter_rows(min_row=2), 0):
        pnl_cell = ws1.cell(i + 2, 4)
        v = pnl_cell.value or 0
        if v > 0:
            pnl_cell.font = Font(color='1D9E75', bold=True)
        elif v < 0:
            pnl_cell.font = Font(color='E24B4A', bold=True)

    for col, w in enumerate([22, 14, 14, 12, 10, 14, 30], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # Sheet 2: Wallets
    ws2 = wb.create_sheet('Кошельки')
    style_header(ws2, ['Адрес', 'Протоколы', 'Метка', 'Добавлен'])
    for w in db.get_wallets():
        protos = ', '.join(w.get('protocols') or []) or '—'
        ws2.append([w['address'], protos, w.get('label', ''), w['added_at']])
    style_rows(ws2)
    for col, width in enumerate([18, 45, 20, 22], 1):
        ws2.column_dimensions[get_column_letter(col)].width = width

    # Sheet 3: Summary
    ws3 = wb.create_sheet('Сводка')
    style_header(ws3, ['Показатель', 'Значение'])
    s = db.get_stats()
    pnl = round(s['total_earned'] - s['total_spent'], 2)
    for row in [
        ['Всего протоколов', s['total']],
        ['Активных', s['active']],
        ['Кошельков', s['total_wallets']],
        ['Потрачено ($)', s['total_spent']],
        ['Заработано ($)', s['total_earned']],
        ['P&L ($)', pnl],
        ['Дата экспорта', datetime.now().strftime('%d.%m.%Y %H:%M')],
    ]:
        ws3.append(row)
    style_rows(ws3)
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 20

    # Save to Downloads
    downloads = os.path.join(os.path.expanduser('~'), 'Downloads')
    os.makedirs(downloads, exist_ok=True)
    filename = f'farmtrack_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    filepath = os.path.join(downloads, filename)
    wb.save(filepath)

    return jsonify({'ok': True, 'filename': filename, 'path': filepath, 'folder': downloads})


# --- Telegram ---

@app.route('/api/telegram/config', methods=['GET'])
def get_telegram_config():
    cfg = db.get_telegram_config()
    token = cfg.get('bot_token', '')
    if token:
        cfg['bot_token'] = '***' + token[-4:]
    return jsonify(cfg)


@app.route('/api/telegram/config', methods=['POST'])
def set_telegram_config():
    global _tg_monitor
    data = request.json or {}
    fields = {}
    for key in ('chat_id', 'alert_threshold_pct', 'alert_cooldown_minutes',
                'report_interval_minutes', 'check_interval_minutes',
                'alerts_enabled', 'reports_enabled', 'enabled'):
        if key in data:
            fields[key] = data[key]
    # Only update token if user sent a real value (not masked)
    token = (data.get('bot_token') or '').strip()
    if token and not token.startswith('***'):
        fields['bot_token'] = token

    if fields:
        db.set_telegram_config(**fields)

    cfg = db.get_telegram_config()
    if _tg_monitor:
        _tg_monitor.update_config(cfg)
        if cfg.get('enabled') and cfg.get('bot_token') and cfg.get('chat_id'):
            if not _tg_monitor.is_running():
                _tg_monitor.start()
        else:
            _tg_monitor.stop()

    return jsonify({'ok': True})


@app.route('/api/telegram/test', methods=['POST'])
def test_telegram():
    cfg = db.get_telegram_config()
    if not cfg.get('bot_token') or not cfg.get('chat_id'):
        return jsonify({'error': 'bot_token и chat_id не заполнены'}), 400
    ok = tg.send_message(cfg['bot_token'], cfg['chat_id'], '✅ FarmTrack подключён!')
    if ok:
        return jsonify({'ok': True})
    return jsonify({'error': 'Не удалось отправить сообщение. Проверьте токен и chat_id.'}), 502
